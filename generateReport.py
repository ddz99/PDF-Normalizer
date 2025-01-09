import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Define the folder containing the JSON files
input_folder = "Extracted_JSON"

# Function to process JSON files
def process_json_files(folder):
    all_data = {}

    for file_name in os.listdir(folder):
        if file_name.endswith(".json"):
            file_path = os.path.join(folder, file_name)
            with open(file_path, "r") as file:
                json_data = json.load(file)

                for key, value in json_data.items():
                    if key not in all_data:
                        all_data[key] = {}

                    # Add manufacturer prices as columns
                    for manufacturer, price in value.items():
                        all_data[key][manufacturer] = price

    return all_data

# Process JSON files
processed_data = process_json_files(input_folder)

# Convert to a DataFrame
df = pd.DataFrame.from_dict(processed_data, orient="index").reset_index()
df.columns = ["Key"] + list(df.columns[1:])  # Rename the first column to "Key"

# Enhance the Excel file with formatting
wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Add headers with styling
headers = list(df.columns)
for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Add data rows
for row_num, row_data in enumerate(df.values, start=2):
    for col_num, value in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_num, value=value)

# Save the styled Excel file
styled_excel_file = "report.xlsx"
wb.save(styled_excel_file)

print(f"Report generated: {styled_excel_file}")